import os 
import pandas as pd 
import numpy as np 
import matplotlib.pyplot as plt
import cv2 
import easyocr
from openpyxl import Workbook, load_workbook
from ast import literal_eval
from openpyxl.utils import get_column_letter

def read_image(image_path):
    image = cv2.imread(image_path)
    # plt.imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))  #Convert BGR to RGB for correct display
    # plt.show()
    return image

def rotate_image(image):
    # Get the height and width of the image
    height, width = image.shape[:2]
    
    # Define the rotation angle (in degrees)
    angle = -0.1 # Adjust this value as needed, e.g., 5 to 10 degrees

    # Calculate the rotation matrix
    rotation_matrix = cv2.getRotationMatrix2D((width / 2, height / 2), angle, 1.0)

    # Apply the rotation to the image
    rotated_image = cv2.warpAffine(image, rotation_matrix, (width, height), borderMode=cv2.BORDER_REFLECT)

    # plt.imshow(cv2.cvtColor(rotated_image, cv2.COLOR_BGR2RGB))
    # plt.show()
    return rotated_image

def preprocess(rotated_image):
    gray = cv2.cvtColor(rotated_image, cv2.COLOR_BGR2GRAY)
    # blurred = cv2.GaussianBlur(gray, (5, 5), 0)\
        
    edges = cv2.Canny(gray, 50, 150)
    return edges

def table_crop(edges):
    contours, _ = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    
    # Filter contours based on area (adjust the threshold accordingly)
    valid_contours = [contour for contour in contours if cv2.contourArea(contour) > 4000]
    
    if not valid_contours:
        return None  # No valid contours found
    
    # Combine bounding boxes of all valid contours
    combined_bbox = cv2.boundingRect(np.concatenate(valid_contours))
    
    x, y, w, h = combined_bbox
    cropped_image = rotated_image[y:y+h, x:x+w]
    
    # print(cropped_image.shape)
    return cropped_image

def table_preprocess(cropped_image):
    threshold = 225
    _, binary_image = cv2.threshold(cropped_image, threshold, 255, cv2.THRESH_BINARY)
    # plt.imshow(binary_image)  # Specify colormap for binary image
    # plt.show()

    v_kernel = np.ones((5, 1), np.uint8)
    dilate_image_v = cv2.dilate(binary_image, v_kernel, iterations=16)
    vertical_lines = cv2.erode(dilate_image_v, v_kernel, iterations=25)
    # plt.imshow(vertical_lines)
    # plt.show()

    h_kernel = np.ones((1, 5), np.uint8)  # New horizontal kernel
    dilate_image_h = cv2.dilate(binary_image, h_kernel, iterations=15)
    horizontal_lines = cv2.erode(dilate_image_h, h_kernel, iterations=25)
    # plt.imshow(horizontal_lines)
    # plt.show()
    
    return vertical_lines, horizontal_lines

# Find column boundaries
def vertical_column_ocr(vertical_lines, horizontal_lines, cropped_image, excel_file):
    column_boundaries = []
    row_boundaries = []
    in_column = False
    in_row = False

    # Create a new workbook and worksheet outside of the loop
    wb = Workbook()
    ws = wb.active 
    
    # Find column boundaries
    for i in range(vertical_lines.shape[1]):
        if not in_column and 255 in vertical_lines[:, i]:
            in_column = True
            start_column = i
        elif in_column and (i == vertical_lines.shape[1] - 1 or not 255 in vertical_lines[:, i]):
            in_column = False
            end_column = i
            column_boundaries.append((start_column, end_column))

    # Find row boundaries and count rows per column
    for i in range(horizontal_lines.shape[0]):
        if not in_row and 255 in horizontal_lines[i, :]:
            in_row = True
            start_row = i
        elif in_row and (i == horizontal_lines.shape[0] - 1 or not 255 in horizontal_lines[i, :]):
            in_row = False
            end_row = i
            row_boundaries.append((start_row, end_row))


    # Process and save columns and rows
    for index, (start_col, end_col) in enumerate(column_boundaries, start=1):
        current_row_index = 1  # Initialize the row index for each column

        for start_row, end_row in row_boundaries:
            # Crop the cell
            cell = cropped_image[start_row:end_row, start_col:end_col]

            # Your processing steps for the cell
            # For example, let's apply a simple thresholding to the 
            threshold = 210
            _, processed_cell = cv2.threshold(cell, threshold, 255, cv2.THRESH_BINARY)
            # plt.imshow(processed_cell)
            # plt.show()
            # Optionally, remove the processed cell from the original image
            cropped_image[start_row:end_row, start_col:end_col] = 255  # Assuming grayscale, change to 255

            reader = easyocr.Reader(['en'])  # You can specify the language(s) you want to use

            results = reader.readtext(processed_cell, detail=0)
            str1 = ''
            for i in results:
                str1 = str1 + ' ' + i

            print(str1)
            # print(results)

            # Check if results list contains only non-empty elements
            if results and any(str1 for str1 in results):
                # Append the concatenated result to the next row in the existing worksheet
                ws.cell(row=current_row_index, column=index, value="\n".join(results))

            # Increment row index for each data point in the column
            current_row_index += 1

    wb.save(excel_file)

if __name__ == "__main__":
    image_path = 'OCR/images/phptNTgqD.jpg'
    excel_file = 'OCR/examples/phptNTgqD.xlsx'
    
    image = read_image(image_path)
    rotated_image = rotate_image(image)
    edges = preprocess(rotated_image)
    cropped_image = table_crop(edges)
    vertical_lines, horizontal_lines = table_preprocess(cropped_image)
    vertical_column_ocr(vertical_lines, horizontal_lines, cropped_image, excel_file)
